import customtkinter as ctk
import tkinter as tk
from tkinter import filedialog, messagebox
import openpyxl
import os
import sys
import logging
import threading
import unicodedata
import subprocess
from datetime import datetime
from typing import Any, Callable
import updater
from styles import (
    AZUL, AZUL_MED, AZUL_CLARO, ACENTO, VERDE, ROJO,
    GRIS_BG, GRIS_BORDE, BLANCO, TEXTO, TEXTO_SUAVE, PASOS,
    actualizar_para_tema,
)
from utils import reformatear_fecha, copiar_a_temp, detectar_campos_word, limpiar_nombre
from docx import Document
from engine import generar_documentos, PDF_DISPONIBLE

DND_DISPONIBLE: bool = False
try:
    import tkinterdnd2
    DND_DISPONIBLE = True
except ImportError:
    pass
from config import cargar, guardar


def _iniciar_log() -> str:
    if getattr(sys, "frozen", False):
        base = os.path.join(os.environ.get("APPDATA", ""), "GeneradorDocumentos")
        os.makedirs(base, exist_ok=True)
        log_path = os.path.join(base, "app.log")
    else:
        log_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "app.log")

    logging.basicConfig(
        filename=log_path,
        level=logging.INFO,
        format="%(asctime)s [%(levelname)s] %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
        encoding="utf-8",
    )
    return log_path


_iniciar_log()
log = logging.getLogger(__name__)


def _normalizar(texto: str) -> str:
    texto = unicodedata.normalize("NFKD", texto).encode("ascii", "ignore").decode("ascii")
    return texto.lower().strip()



class AppContratos(ctk.CTk):
    def __init__(self) -> None:
        super().__init__()
        self.title("Generador de Documentos")
        self.config_data = cargar()
        ctk.set_appearance_mode(self.config_data.get("apariencia", "light"))
        actualizar_para_tema()
        global AZUL_CLARO, GRIS_BG, TEXTO, TEXTO_SUAVE, GRIS_BORDE
        import styles as _st
        AZUL_CLARO = _st.AZUL_CLARO
        GRIS_BG = _st.GRIS_BG
        TEXTO = _st.TEXTO
        TEXTO_SUAVE = _st.TEXTO_SUAVE
        GRIS_BORDE = _st.GRIS_BORDE
        ancho = self.config_data.get("ancho", 780)
        alto = self.config_data.get("alto", 700)
        self.geometry(f"{ancho}x{alto}")
        self.minsize(720, 640)
        self.resizable(True, True)
        self.configure(fg_color=GRIS_BG)

        self.excel_path: tk.StringVar = tk.StringVar(value="")
        self.word_path: tk.StringVar = tk.StringVar(value="")
        self.personas: list[dict[str, Any]] = []
        self.columnas: list[str] = []
        self.persona_sel: dict[str, Any] | None = None
        self.campos_extra: dict[str, Any] = {}

        self._build_ui()
        self.bind("<Control-o>", lambda e: self._browse(
            [("Excel", "*.xlsx *.xls")], self._cargar_excel, self.excel_path,
        ))
        self.bind("<Control-w>", lambda e: self._browse(
            [("Word", "*.docx")], self._validar_word, self.word_path,
        ))
        self.bind("<Control-g>", lambda e: self._generar())
        self.bind("<Control-r>", lambda e: self._reiniciar())
        self.protocol("WM_DELETE_WINDOW", self._on_cerrar)

    def _on_cerrar(self) -> None:
        if messagebox.askokcancel("Salir", "¿Estás seguro de que deseas salir?"):
            self._guardar_config()
            log.info("Cerrando aplicación por solicitud del usuario")
            self.destroy()

    def _verificar_actualizacion(self) -> None:
        updater.verificar_actualizacion(
            on_update_available=self._on_actualizacion_disponible,
            on_error=None,
        )

    def _on_actualizacion_disponible(self, version: str, url: str, sha256: str = "") -> None:
        self.after(0, lambda: self._mostrar_dialogo_actualizacion(version, url, sha256))

    def _mostrar_dialogo_actualizacion(self, version: str, url: str, sha256: str = "") -> None:
        respuesta = messagebox.askyesno(
            "Actualización disponible",
            f"Hay una nueva versión disponible: v{version}\n\n"
            f"¿Deseas actualizar ahora?\n"
            f"(La aplicación se reiniciará automáticamente)",
        )
        if respuesta:
            self._iniciar_descarga(url, sha256)

    def _iniciar_descarga(self, url: str, sha256: str = "") -> None:
        ventana = ctk.CTkToplevel(self)
        ventana.title("Actualizando…")
        ventana.geometry("360x140")
        ventana.resizable(False, False)
        ventana.grab_set()

        ctk.CTkLabel(
            ventana, text="Descargando actualización…",
            font=ctk.CTkFont(size=14, weight="bold"),
        ).pack(pady=(20, 10))

        barra = ctk.CTkProgressBar(ventana, width=300)
        barra.set(0)
        barra.pack(pady=4)

        lbl_pct = ctk.CTkLabel(ventana, text="0%", font=ctk.CTkFont(size=12))
        lbl_pct.pack()

        def on_progreso(pct: int) -> None:
            self.after(0, lambda p=pct: (barra.set(p / 100), lbl_pct.configure(text=f"{p}%")))

        def on_error(msg: str) -> None:
            self.after(0, lambda: (
                ventana.grab_release(), ventana.destroy(),
                messagebox.showerror("Error de actualización",
                    f"No se pudo descargar la actualización:\n{msg}"),
            ))

        updater.descargar_e_instalar(url, sha256, on_progreso=on_progreso, on_error=on_error)

    # ── UI ────────────────────────────────────────────────────────────────────
    def _build_ui(self) -> None:
        header = ctk.CTkFrame(self, fg_color=AZUL, corner_radius=0, height=80)
        header.pack(fill="x")
        header.pack_propagate(False)

        ctk.CTkLabel(
            header,
            text="Generador de Documentos",
            font=ctk.CTkFont(family="Georgia", size=22, weight="bold"),
            text_color=BLANCO,
        ).pack(side="left", padx=28, pady=20)

        version_local = updater.leer_version_local()
        ctk.CTkLabel(
            header,
            text=f"Contratos · Finiquitos · Cartas   |   v{version_local}",
            font=ctk.CTkFont(size=12),
            text_color="#A8C4E8",
        ).pack(side="right", padx=28, pady=20)

        self.frame_barra = ctk.CTkFrame(self, fg_color=AZUL_CLARO, corner_radius=0, height=44)
        self.frame_barra.pack(fill="x")
        self.frame_barra.pack_propagate(False)
        self._build_barra_pasos()

        self.frame_contenido = ctk.CTkScrollableFrame(self, fg_color=GRIS_BG, corner_radius=0)
        self.frame_contenido.pack(fill="both", expand=True, padx=24, pady=20)

        self.nav = ctk.CTkFrame(self, fg_color=BLANCO, corner_radius=0, height=60)
        self.nav.pack(fill="x", side="bottom")
        self.nav.pack_propagate(False)

        self.btn_oscuro = ctk.CTkButton(
            self.nav, text="🌙", font=ctk.CTkFont(size=16), width=38, height=38,
            fg_color="#1E2A3A", hover_color="#2D4A7A",
            corner_radius=8, command=self._alternar_apariencia,
        )
        self.btn_oscuro.pack(side="right", padx=(0, 6), pady=10)
        self._actualizar_boton_oscuro()

        self.btn_actualizar = ctk.CTkButton(
            self.nav, text="Buscar actualizaciones", font=ctk.CTkFont(size=11),
            fg_color="transparent", hover_color=AZUL_CLARO, text_color=TEXTO_SUAVE,
            corner_radius=8, height=32, command=self._verificar_actualizacion,
        )
        self.btn_actualizar.pack(side="right", padx=(0, 4), pady=10)

        self.btn_inicio = ctk.CTkButton(
            self.nav, text="Reiniciar", font=ctk.CTkFont(size=12),
            fg_color="transparent", hover_color=AZUL_CLARO, text_color=TEXTO_SUAVE,
            corner_radius=8, height=38, width=100, command=self._reiniciar,
        )
        self.btn_inicio.pack(side="left", padx=(6, 0), pady=10)

        self.btn_atras = ctk.CTkButton(
            self.nav, text="Atras", font=ctk.CTkFont(size=13, weight="bold"),
            fg_color=GRIS_BORDE, hover_color="#B0BCC8", text_color=TEXTO,
            corner_radius=8, height=38, width=120, command=self._paso_atras,
        )
        self.btn_atras.pack(side="left", padx=20, pady=10)

        self.btn_siguiente = ctk.CTkButton(
            self.nav, text="Siguiente", font=ctk.CTkFont(size=13, weight="bold"),
            fg_color=AZUL_MED, hover_color=AZUL, text_color=BLANCO,
            corner_radius=8, height=38, width=120, command=self._paso_siguiente,
        )
        self.btn_siguiente.pack(side="right", padx=20, pady=10)

        self.paso_actual = 0
        self._renderizar_paso()
        self._actualizar_colores()

    def _build_barra_pasos(self) -> None:
        titulos = ["1. Archivos", "2. Trabajadores", "3. Datos faltantes", "4. Generar"]
        for w in self.frame_barra.winfo_children():
            w.destroy()
        frame_fg = self._frame_fg_theme()
        for i, titulo in enumerate(titulos):
            activo = i == getattr(self, "paso_actual", 0)
            ctk.CTkLabel(
                self.frame_barra, text=titulo,
                font=ctk.CTkFont(size=12, weight="bold" if activo else "normal"),
                text_color=AZUL if activo else TEXTO_SUAVE,
                fg_color=frame_fg if activo else "transparent",
                corner_radius=6, padx=12, pady=4,
            ).pack(side="left", padx=6, pady=8)

    def _renderizar_paso(self) -> None:
        self._limpiar_variables()
        for w in self.frame_contenido.winfo_children():
            w.destroy()
        self._build_barra_pasos()

        pasos_fn: list[Callable[[], None]] = [
            self._paso_archivos,
            self._paso_trabajadores,
            self._paso_datos_faltantes,
            self._paso_generar,
        ]
        pasos_fn[self.paso_actual]()

        self.btn_atras.configure(state="normal" if self.paso_actual > 0 else "disabled")
        ultimo = self.paso_actual == len(PASOS) - 1
        self.btn_siguiente.configure(
            text="Generar" if ultimo else "Siguiente",
            fg_color=VERDE if ultimo else AZUL_MED,
            hover_color="#15803D" if ultimo else AZUL,
            command=self._generar if ultimo else self._paso_siguiente,
        )

    def _paso_siguiente(self) -> None:
        if self.paso_actual == 0:
            if not self.excel_path.get() or not self.word_path.get():
                messagebox.showwarning("Atencion", "Selecciona ambos archivos antes de continuar.")
                return
        if self.paso_actual == 1:
            if not getattr(self, "trabajadores_sel", []):
                messagebox.showwarning("Atencion", "Selecciona al menos un trabajador.")
                return

            campos_word = self._detectar_campos_word()
            campos_excel = {k.lower() for k in self.columnas}
            campos_en_ambos = [c for c in campos_word if c.lower() in campos_excel]

            avisos: list[str] = []
            for persona in self.trabajadores_sel:
                col_id = getattr(self, "col_identificador_var", tk.StringVar(value=self.columnas[0])).get()
                nombre = str(persona.get(col_id, str(list(persona.values())[0])))
                faltantes = [
                    c for c in campos_en_ambos
                    if not persona.get(c) and not persona.get(c.lower())
                ]
                if faltantes:
                    avisos.append(f"{nombre}: falta {', '.join(faltantes)}")

            if avisos:
                msg = "Los siguientes trabajadores tienen datos vacios:\n\n" + "\n".join(avisos)
                msg += "\n\nPuedes continuar de todas formas o volver a revisar."
                continuar = messagebox.askyesno("Datos incompletos", msg)
                if not continuar:
                    return
        if self.paso_actual < len(PASOS) - 1:
            self.paso_actual += 1
            self._renderizar_paso()

    def _paso_atras(self) -> None:
        if self.paso_actual > 0:
            self.paso_actual -= 1
            self._renderizar_paso()

    def _alternar_apariencia(self) -> None:
        modo = "dark" if ctk.get_appearance_mode() == "Light" else "light"
        ctk.set_appearance_mode(modo)
        actualizar_para_tema()
        global AZUL_CLARO, GRIS_BG, TEXTO, TEXTO_SUAVE, GRIS_BORDE
        import styles as _st
        AZUL_CLARO = _st.AZUL_CLARO
        GRIS_BG = _st.GRIS_BG
        TEXTO = _st.TEXTO
        TEXTO_SUAVE = _st.TEXTO_SUAVE
        GRIS_BORDE = _st.GRIS_BORDE
        self._actualizar_boton_oscuro()
        self._actualizar_colores()
        self._renderizar_paso()
        self.config_data["apariencia"] = modo

    def _frame_fg_theme(self) -> str:
        raw = ctk.ThemeManager.theme.get("CTkFrame", {}).get("fg_color", ["#FFFFFF", "#2B2B2B"])
        modo = ctk.AppearanceModeTracker.get_mode()
        return raw[modo] if isinstance(raw, (list, tuple)) else raw

    def _actualizar_colores(self) -> None:
        frame_fg = self._frame_fg_theme()
        self.configure(fg_color=GRIS_BG)
        self.frame_contenido.configure(fg_color=GRIS_BG)
        self.frame_barra.configure(fg_color=AZUL_CLARO)
        for w in self.frame_barra.winfo_children():
            if isinstance(w, ctk.CTkLabel):
                try:
                    idx = self.frame_barra.winfo_children().index(w)
                    activo = idx == getattr(self, "paso_actual", 0)
                    w.configure(
                        text_color=AZUL if activo else TEXTO_SUAVE,
                        fg_color=frame_fg if activo else "transparent",
                    )
                except ValueError:
                    pass
        for btn in [self.btn_inicio, self.btn_actualizar]:
            btn.configure(hover_color=AZUL_CLARO, text_color=TEXTO_SUAVE)
        self.btn_oscuro.configure(fg_color="#1E2A3A", hover_color="#2D4A7A")
        self.btn_atras.configure(
            fg_color=GRIS_BORDE, hover_color="#B0BCC8", text_color=TEXTO,
        )
        self.btn_siguiente.configure(
            fg_color=AZUL_MED, hover_color=AZUL, text_color=BLANCO,
        )
        if hasattr(self, "nav"):
            self.nav.configure(fg_color=frame_fg)

    def _actualizar_boton_oscuro(self) -> None:
        es_oscuro = ctk.get_appearance_mode() == "Dark"
        self.btn_oscuro.configure(text="☀️" if es_oscuro else "🌙")

    def _reiniciar(self) -> None:
        if self.paso_actual > 0 and not messagebox.askokcancel(
            "Reiniciar", "Se perderán los datos cargados. ¿Continuar?"
        ):
            return
        self.personas = []
        self.columnas = []
        self.trabajadores_sel = []
        self.campos_extra_vars = {}
        self.paso_actual = 0
        if hasattr(self, "col_identificador_var"):
            del self.col_identificador_var
        if hasattr(self, "col_vars"):
            del self.col_vars
        if hasattr(self, "espacios_var"):
            del self.espacios_var
        for var in [self.excel_path, self.word_path]:
            try:
                for modo, cbname in var.trace_info():
                    var.trace_remove(modo, cbname)
            except AttributeError:
                pass
        self._renderizar_paso()
        self.excel_path.set("")
        self.word_path.set("")

    def _limpiar_variables(self) -> None:
        if hasattr(self, "busqueda_var") and hasattr(self, "_trace_id_busqueda"):
            try:
                self.busqueda_var.trace_remove("write", self._trace_id_busqueda)
            except Exception:
                pass
            del self.busqueda_var
            del self._trace_id_busqueda

        if hasattr(self, "check_vars_trabajadores"):
            del self.check_vars_trabajadores

    # ── PASO 1 – ARCHIVOS ─────────────────────────────────────────────────────
    def _paso_archivos(self) -> None:
        parent = self.frame_contenido
        ctk.CTkLabel(
            parent,
            text="Selecciona los archivos",
            font=ctk.CTkFont(size=16, weight="bold"),
            text_color=TEXTO,
        ).pack(anchor="w", pady=(0, 16))

        self._file_row(
            parent, "Archivo Excel con personas:",
            "Seleccionar Excel (.xlsx)", self.excel_path,
            self._cargar_excel, [("Excel", "*.xlsx *.xls")],
        )
        self._file_row(
            parent, "Plantilla Word (base del documento):",
            "Seleccionar Word (.docx)", self.word_path,
            self._validar_word, [("Word", "*.docx")],
        )

        # Selector de hoja (se muestra al cargar Excel con múltiples hojas)
        self.frame_hoja = ctk.CTkFrame(parent, fg_color="transparent")
        self.hoja_label = ctk.CTkLabel(
            self.frame_hoja, text="Hoja del Excel:",
            font=ctk.CTkFont(size=13, weight="bold"), text_color=TEXTO,
        )
        self.hoja_label.pack(side="left", padx=(0, 10))
        self.sheet_var = tk.StringVar(value="")
        self.hoja_combo = ctk.CTkOptionMenu(
            self.frame_hoja, variable=self.sheet_var,
            values=[""], command=self._on_hoja_cambiada,
            font=ctk.CTkFont(size=12),
            fg_color=AZUL_CLARO, text_color=TEXTO,
            button_color=AZUL_MED, button_hover_color=AZUL,
            dynamic_resizing=False,
        )
        self.hoja_combo.pack(side="left")

        self._build_drop_zone(parent)

    def _build_drop_zone(self, parent: ctk.CTkBaseClass) -> None:
        drop_frame = ctk.CTkFrame(
            parent, fg_color="transparent", border_width=2,
            border_color=GRIS_BORDE, corner_radius=8, height=80,
        )
        drop_frame.pack(fill="x", pady=(16, 0))
        drop_frame.pack_propagate(False)

        txt = "Arrastra archivos aqui (*.xlsx, *.docx)" if DND_DISPONIBLE else \
              "Haz clic aqui o usa los botones de arriba para seleccionar archivos"
        lbl = ctk.CTkLabel(
            drop_frame, text=txt,
            font=ctk.CTkFont(size=12), text_color=TEXTO_SUAVE,
        )
        lbl.pack(expand=True)
        lbl.bind("<Button-1>", lambda e: self._browse(
            [("Archivos", "*.xlsx *.xls *.docx")], self._manejar_drop, self.excel_path,
        ))

        if DND_DISPONIBLE:
            self._registrar_drag_drop()

    def _registrar_drag_drop(self) -> None:
        try:
            import tkinterdnd2 as dnd
            dnd.TkinterDnD.require(self)
            self.drop_target_register("*")
            self.dnd_bind("<<Drop>>", self._on_drop)
        except Exception:
            pass

    def _on_drop(self, event: object) -> None:
        archivos = getattr(event, "data", "")
        if not archivos:
            return
        for archivo in archivos.split():
            archivo = archivo.strip("{}")
            if archivo.endswith((".xlsx", ".xls")):
                self.excel_path.set(archivo)
                self._cargar_excel(archivo)
            elif archivo.endswith(".docx"):
                self.word_path.set(archivo)
                self._validar_word(archivo)

    def _manejar_drop(self, path: str) -> None:
        if path.endswith((".xlsx", ".xls")):
            self.excel_path.set(path)
            self._cargar_excel(path)
        elif path.endswith(".docx"):
            self.word_path.set(path)
            self._validar_word(path)

    def _file_row(
        self,
        parent: ctk.CTkBaseClass,
        label_txt: str,
        btn_txt: str,
        var: tk.StringVar,
        cmd: Callable[[str], None],
        filetypes: list[tuple[str, str]],
    ) -> None:
        ctk.CTkLabel(
            parent, text=label_txt,
            font=ctk.CTkFont(size=13, weight="bold"),
            text_color=TEXTO, anchor="w",
        ).pack(fill="x", pady=(6, 4))

        row = ctk.CTkFrame(parent, fg_color="transparent")
        row.pack(fill="x", pady=(0, 10))

        ctk.CTkEntry(
            row, textvariable=var, state="readonly",
            placeholder_text="Ningún archivo seleccionado",
            font=ctk.CTkFont(size=12), fg_color=GRIS_BG,
            border_color=GRIS_BORDE, text_color=TEXTO_SUAVE, height=38,
        ).pack(side="left", fill="x", expand=True, padx=(0, 10))

        ctk.CTkButton(
            row, text=btn_txt, font=ctk.CTkFont(size=12, weight="bold"),
            fg_color=AZUL_MED, hover_color=AZUL, text_color=BLANCO,
            corner_radius=8, height=38,
            command=lambda ft=filetypes, c=cmd, v=var: self._browse(ft, c, v),
        ).pack(side="right")

    def _browse(
        self,
        filetypes: list[tuple[str, str]],
        callback: Callable[[str], None],
        var: tk.StringVar,
    ) -> None:
        path = filedialog.askopenfilename(filetypes=filetypes)
        if path:
            var.set(path)
            callback(path)

    # ── PASO 2 – TRABAJADORES ─────────────────────────────────────────────────
    def _paso_trabajadores(self) -> None:
        parent = self.frame_contenido

        ctk.CTkLabel(
            parent, text="Selecciona los trabajadores",
            font=ctk.CTkFont(size=16, weight="bold"), text_color=TEXTO,
        ).pack(anchor="w", pady=(0, 4))

        ctk.CTkLabel(
            parent, text="Puedes seleccionar uno o varios.",
            font=ctk.CTkFont(size=11), text_color=TEXTO_SUAVE,
        ).pack(anchor="w", pady=(0, 10))

        # Selector de columnas visibles
        ctk.CTkLabel(
            parent, text="Columnas a mostrar en la lista:",
            font=ctk.CTkFont(size=12, weight="bold"), text_color=TEXTO,
        ).pack(anchor="w", pady=(0, 4))

        frame_cols = ctk.CTkFrame(parent, fg_color=AZUL_CLARO, corner_radius=8)
        frame_cols.pack(fill="x", pady=(0, 10))

        def _es_col_identificadora(col: str) -> bool:
            c = col.lower()
            return any(k in c for k in ["rut", "run", "nombre", "nombres", "apellido"])

        if not hasattr(self, "col_vars"):
            self.col_vars = {
                col: tk.BooleanVar(value=_es_col_identificadora(col))
                for col in self.columnas
            }

        max_por_fila = 4
        for i, col in enumerate(self.columnas):
            if i % max_por_fila == 0:
                fila_cols = ctk.CTkFrame(frame_cols, fg_color="transparent")
                fila_cols.pack(fill="x", padx=8, pady=2)

            ctk.CTkCheckBox(
                fila_cols, text=col, variable=self.col_vars[col],
                font=ctk.CTkFont(size=11),
                command=self._refrescar_lista_trabajadores,
            ).pack(side="left", padx=10, pady=4)

        ctk.CTkLabel(
            parent,
            text="Columna para identificar a cada persona (se usa como nombre en datos faltantes):",
            font=ctk.CTkFont(size=12, weight="bold"), text_color=TEXTO,
        ).pack(anchor="w", pady=(10, 4))

        if not hasattr(self, "col_identificador_var"):
            saved = self.config_data.get("col_identificador", "")
            default = saved if saved in self.columnas else (
                next((c for c in self.columnas if any(k in c.lower() for k in ["rut", "run"])), None)
                or next((c for c in self.columnas if any(k in c.lower() for k in ["nombre", "nombres"])), None)
                or self.columnas[0]
            )
            self.col_identificador_var = tk.StringVar(value=default)
        frame_id = ctk.CTkFrame(parent, fg_color=AZUL_CLARO, corner_radius=8)
        frame_id.pack(fill="x", pady=(0, 10))

        for i, col in enumerate(self.columnas):
            if i % max_por_fila == 0:
                fila_id = ctk.CTkFrame(frame_id, fg_color="transparent")
                fila_id.pack(fill="x", padx=8, pady=2)

            ctk.CTkRadioButton(
                fila_id, text=col, variable=self.col_identificador_var,
                value=col, font=ctk.CTkFont(size=11), text_color=TEXTO,
            ).pack(side="left", padx=10, pady=4)

        # Búsqueda
        self.busqueda_var = tk.StringVar()
        self._trace_id_busqueda = self.busqueda_var.trace_add("write", lambda *_: self._refrescar_lista_trabajadores())

        ctk.CTkEntry(
            parent, textvariable=self.busqueda_var,
            placeholder_text="Buscar...", font=ctk.CTkFont(size=13),
            height=40, border_color=ACENTO,
        ).pack(fill="x", pady=(0, 8))

        # Botones seleccionar/deseleccionar todos
        fila_sel = ctk.CTkFrame(parent, fg_color="transparent")
        fila_sel.pack(fill="x", pady=(0, 6))

        ctk.CTkButton(
            fila_sel, text="Seleccionar todos", font=ctk.CTkFont(size=11),
            fg_color=AZUL_CLARO, text_color=TEXTO, hover_color=GRIS_BORDE,
            height=28, corner_radius=6, command=lambda: self._sel_todos(True),
        ).pack(side="left", padx=(0, 6))

        ctk.CTkButton(
            fila_sel, text="Deseleccionar todos", font=ctk.CTkFont(size=11),
            fg_color=AZUL_CLARO, text_color=TEXTO, hover_color=GRIS_BORDE,
            height=28, corner_radius=6, command=lambda: self._sel_todos(False),
        ).pack(side="left")

        # Lista
        self.lista_frame = ctk.CTkFrame(
            parent, fg_color="transparent", corner_radius=0,
        )
        self.lista_frame.pack(fill="x")

        if not hasattr(self, "trabajadores_sel"):
            self.trabajadores_sel = []

        self._refrescar_lista_trabajadores()

    def _refrescar_lista_trabajadores(self) -> None:
        for w in self.lista_frame.winfo_children():
            w.destroy()

        cols_visibles = [c for c, v in self.col_vars.items() if v.get()]
        if not cols_visibles:
            cols_visibles = self.columnas[:1]

        q = getattr(self, "busqueda_var", tk.StringVar()).get().strip()
        q_norm = _normalizar(q) if q else ""
        personas = [
            p for p in self.personas
            if not q_norm or any(q_norm in _normalizar(str(v)) for v in p.values())
        ]

        if not hasattr(self, "check_vars_trabajadores"):
            self.check_vars_trabajadores = {}

        for p in personas:
            key = str(list(p.values()))
            if key not in self.check_vars_trabajadores:
                self.check_vars_trabajadores[key] = tk.BooleanVar(value=p in self.trabajadores_sel)

            partes = [str(p.get(c, "")) for c in cols_visibles if p.get(c)]
            display = "   |   ".join(partes) if partes else "Sin datos"

            def on_toggle(persona: dict[str, Any] = p, k: str = key) -> None:
                if self.check_vars_trabajadores[k].get():
                    if persona not in self.trabajadores_sel:
                        self.trabajadores_sel.append(persona)
                else:
                    if persona in self.trabajadores_sel:
                        self.trabajadores_sel.remove(persona)

            ctk.CTkCheckBox(
                self.lista_frame, text=display,
                variable=self.check_vars_trabajadores[key],
                font=ctk.CTkFont(size=12), text_color=TEXTO, command=on_toggle,
            ).pack(anchor="w", pady=3, padx=6)

    def _sel_todos(self, valor: bool) -> None:
        for key, var in self.check_vars_trabajadores.items():
            var.set(valor)
        if valor:
            self.trabajadores_sel = list(self.personas)
        else:
            self.trabajadores_sel = []

    # ── PASO 3 – DATOS FALTANTES ──────────────────────────────────────────────
    def _paso_datos_faltantes(self) -> None:
        parent = self.frame_contenido

        ctk.CTkLabel(
            parent, text="Datos faltantes",
            font=ctk.CTkFont(size=16, weight="bold"), text_color=TEXTO,
        ).pack(anchor="w", pady=(0, 4))

        campos_word = self._detectar_campos_word()
        campos_excel = {k.lower() for k in self.columnas}
        campos_faltantes_word = [c for c in campos_word if c.lower() not in campos_excel]

        campos_en_ambos = [c for c in campos_word if c.lower() in campos_excel]

        self.campos_extra_vars = {}
        hay_algo = False

        ctk.CTkLabel(
            parent, text="Formato para fechas del Excel:",
            font=ctk.CTkFont(size=13, weight="bold"), text_color=TEXTO,
        ).pack(anchor="w", pady=(0, 4))

        if not hasattr(self, "formato_fecha_var"):
            self.formato_fecha_var = tk.StringVar(
                value=self.config_data.get("formato_fecha", "%d/%m/%Y")
            )

        formatos_globales: list[tuple[str, str]] = [
            ("13/05/2025",          "%d/%m/%Y"),
            ("13-05-2025",          "%d-%m-%Y"),
            ("13 de Mayo del 2025", "texto"),
            ("2025-05-13",          "%Y-%m-%d"),
        ]

        fila_fmt = ctk.CTkFrame(parent, fg_color="transparent")
        fila_fmt.pack(fill="x", pady=(0, 14))

        for etiqueta, fmt in formatos_globales:
            ctk.CTkRadioButton(
                fila_fmt, text=etiqueta, variable=self.formato_fecha_var,
                value=fmt, font=ctk.CTkFont(size=11), text_color=TEXTO,
            ).pack(side="left", padx=8)

        for persona in self.trabajadores_sel:
            col_id = getattr(self, "col_identificador_var", tk.StringVar(value=self.columnas[0])).get()
            nombre_persona = str(persona.get(col_id, str(list(persona.values())[0])))
            self.campos_extra_vars[nombre_persona] = {}

            faltantes_excel = [
                c for c in campos_en_ambos
                if not persona.get(c) and not persona.get(c.lower())
            ]

            campos_este_trabajador = campos_faltantes_word + faltantes_excel

            if not campos_este_trabajador:
                continue

            hay_algo = True

            sep = ctk.CTkFrame(parent, fg_color=AZUL_CLARO, corner_radius=8)
            sep.pack(fill="x", pady=(10, 6))
            ctk.CTkLabel(
                sep, text=nombre_persona,
                font=ctk.CTkFont(size=13, weight="bold"), text_color=TEXTO,
            ).pack(anchor="w", padx=12, pady=6)

            for campo in campos_este_trabajador:
                fila = ctk.CTkFrame(parent, fg_color="transparent")
                fila.pack(fill="x", pady=4, padx=8)

                ctk.CTkLabel(
                    fila, text=f"{campo}:",
                    font=ctk.CTkFont(size=12, weight="bold"),
                    text_color=TEXTO, width=180, anchor="w",
                ).pack(side="left")

                if "fecha" in campo.lower():
                    var = tk.StringVar(value=datetime.today().strftime("%d/%m/%Y"))
                    self.campos_extra_vars[nombre_persona][campo] = var

                    contenedor = ctk.CTkFrame(fila, fg_color="transparent")
                    contenedor.pack(side="left", fill="x", expand=True)

                    ctk.CTkEntry(
                        contenedor, textvariable=var,
                        font=ctk.CTkFont(size=12), height=34, border_color=GRIS_BORDE,
                    ).pack(fill="x", pady=(0, 4))

                    fila_formatos = ctk.CTkFrame(contenedor, fg_color="transparent")
                    fila_formatos.pack(fill="x")

                    formatos: list[tuple[str, str]] = [
                        ("13/05/2025",          "%d/%m/%Y"),
                        ("13-05-2025",          "%d-%m-%Y"),
                        ("13 de Mayo del 2025", "texto"),
                        ("2025-05-13",          "%Y-%m-%d"),
                    ]
                    for etiqueta, fmt in formatos:
                        ctk.CTkButton(
                            fila_formatos, text=etiqueta, font=ctk.CTkFont(size=10),
                            fg_color=AZUL_CLARO, text_color=TEXTO, hover_color=GRIS_BORDE,
                            height=24, corner_radius=4,
                            command=lambda f=fmt, variable=var: variable.set(reformatear_fecha(variable.get(), f)),
                        ).pack(side="left", padx=2)
                else:
                    var = tk.StringVar(value="")
                    self.campos_extra_vars[nombre_persona][campo] = var
                    ctk.CTkEntry(
                        fila, textvariable=var, font=ctk.CTkFont(size=12),
                        height=34, border_color=GRIS_BORDE,
                    ).pack(side="left", fill="x", expand=True)

        if not hay_algo:
            ctk.CTkLabel(
                parent,
                text="No hay datos faltantes. Todos los trabajadores tienen sus datos completos.",
                font=ctk.CTkFont(size=13), text_color=VERDE,
            ).pack(anchor="w", pady=10)

    def _detectar_campos_word(self) -> list[str]:
        return detectar_campos_word(self.word_path.get())

    # ── PASO 4 – GENERAR ──────────────────────────────────────────────────────
    def _paso_generar(self) -> None:
        parent = self.frame_contenido

        ctk.CTkLabel(
            parent, text="Configurar y generar",
            font=ctk.CTkFont(size=16, weight="bold"), text_color=TEXTO,
        ).pack(anchor="w", pady=(0, 16))

        ctk.CTkLabel(
            parent, text="Formato del nombre de archivo:",
            font=ctk.CTkFont(size=13, weight="bold"), text_color=TEXTO,
        ).pack(anchor="w", pady=(0, 4))

        if not hasattr(self, "espacios_var"):
            valor_espacios = "guion_bajo" if self.config_data.get("usar_guion_bajo", True) else "espacios"
            self.espacios_var = tk.StringVar(value=valor_espacios)

        fila_espacios = ctk.CTkFrame(parent, fg_color="transparent")
        fila_espacios.pack(fill="x", pady=(0, 12))

        ctk.CTkLabel(
            fila_espacios, text="Espacios en el nombre del archivo:",
            font=ctk.CTkFont(size=12, weight="bold"), text_color=TEXTO,
        ).pack(side="left", padx=(0, 12))

        ctk.CTkRadioButton(
            fila_espacios, text="Usar guion bajo  (mi_documento)",
            variable=self.espacios_var, value="guion_bajo", font=ctk.CTkFont(size=11),
        ).pack(side="left", padx=6)

        ctk.CTkRadioButton(
            fila_espacios, text="Mantener espacios  (mi documento)",
            variable=self.espacios_var, value="espacios", font=ctk.CTkFont(size=11),
        ).pack(side="left", padx=6)

        self.espacios_var.trace_add("write", lambda *_: self._actualizar_preview())

        ctk.CTkLabel(
            parent,
            text="Usa {columna} para insertar datos del trabajador. Ejemplo: {nombre}_{cargo}_{fecha}",
            font=ctk.CTkFont(size=11), text_color=TEXTO_SUAVE,
        ).pack(anchor="w", pady=(0, 6))

        fila_cols = ctk.CTkFrame(parent, fg_color="transparent")
        fila_cols.pack(fill="x", pady=(0, 6))

        default_nombre = self.config_data.get("formato_nombre", "")
        if not default_nombre:
            default_nombre = "{nombre}_{cargo}" if "cargo" in [c.lower() for c in self.columnas] else "{" + self.columnas[0] + "}"
        self.formato_nombre_var = tk.StringVar(value=default_nombre)

        for col in self.columnas[:6]:
            ctk.CTkButton(
                fila_cols, text=f"+{col}", font=ctk.CTkFont(size=10),
                fg_color=AZUL_CLARO, text_color=TEXTO, hover_color=GRIS_BORDE,
                height=24, corner_radius=4,
                command=lambda c=col: self.formato_nombre_var.set(
                    self.formato_nombre_var.get() + f"_{{{c}}}"
                ),
            ).pack(side="left", padx=2)

        ctk.CTkEntry(
            parent, textvariable=self.formato_nombre_var,
            font=ctk.CTkFont(size=12), height=38, border_color=GRIS_BORDE,
        ).pack(fill="x", pady=(0, 6))

        self.lbl_preview = ctk.CTkLabel(
            parent, text="", font=ctk.CTkFont(size=11),
            text_color=TEXTO_SUAVE, anchor="w",
        )
        self.lbl_preview.pack(fill="x", pady=(0, 16))
        self.formato_nombre_var.trace_add("write", lambda *_: self._actualizar_preview())
        self._actualizar_preview()

        n = len(getattr(self, "trabajadores_sel", []))
        ctk.CTkLabel(
            parent, text=f"Se generaran {n} documento(s).",
            font=ctk.CTkFont(size=13), text_color=TEXTO_SUAVE,
        ).pack(anchor="w", pady=(0, 10))

        self.pdf_var = tk.BooleanVar(value=False)
        frame_pdf = ctk.CTkFrame(parent, fg_color="transparent")
        frame_pdf.pack(anchor="w", pady=(0, 10))
        ctk.CTkCheckBox(
            frame_pdf, text="Generar tambien PDF", variable=self.pdf_var,
            font=ctk.CTkFont(size=12), text_color=TEXTO,
            onvalue=True, offvalue=False,
        ).pack(side="left")
        if not PDF_DISPONIBLE:
            ctk.CTkLabel(
                frame_pdf, text="(requiere docx2pdf o LibreOffice)",
                font=ctk.CTkFont(size=10), text_color=TEXTO_SUAVE,
            ).pack(side="left", padx=(6, 0))

        self.progress_bar = ctk.CTkProgressBar(parent, width=400)
        self.progress_bar.set(0)
        self.lbl_progreso = ctk.CTkLabel(parent, text="", font=ctk.CTkFont(size=12), text_color=TEXTO_SUAVE)
        self.lbl_estado = ctk.CTkLabel(parent, text="", font=ctk.CTkFont(size=13), text_color=VERDE)
        self.progress_bar.pack(anchor="w", pady=(0, 4))
        self.progress_bar.pack_forget()
        self.lbl_progreso.pack(anchor="w")
        self.lbl_progreso.pack_forget()
        self.lbl_estado.pack(anchor="w")

    # ── LÓGICA ────────────────────────────────────────────────────────────────
    def _cargar_excel(self, path: str) -> None:
        ruta_lectura: str | None = None
        try:
            ruta_lectura = copiar_a_temp(path)
            wb = openpyxl.load_workbook(ruta_lectura, read_only=True, data_only=True)

            # Configurar selector de hoja
            nombre_hoja = self.sheet_var.get() or wb.sheetnames[0]
            if nombre_hoja not in wb.sheetnames:
                nombre_hoja = wb.sheetnames[0]
            ws = wb[nombre_hoja]

            filas = list(ws.iter_rows(values_only=True))
            if not filas:
                log.warning("Excel vacío seleccionado: %s", path)
                messagebox.showerror("Error", "El Excel está vacío.")
                wb.close()
                return

            headers = [str(c).strip() if c is not None else None for c in filas[0]]
            self.personas = []
            for fila in filas[1:]:
                if any(c is not None for c in fila):
                    persona: dict[str, Any] = {}
                    for i, col in enumerate(headers):
                        if col is not None:
                            valor = str(fila[i]) if i < len(fila) and fila[i] is not None else ""
                            persona[col] = valor
                    self.personas.append(persona)

            self.columnas = [c for c in headers if c is not None]
            sheets_names = wb.sheetnames
            wb.close()

            # Detectar si hay fórmulas (data_only=True da None si no están cacheadas)
            try:
                wb2 = openpyxl.load_workbook(ruta_lectura, read_only=True, data_only=False)
                ws2 = wb2[nombre_hoja]
                filas2 = list(ws2.iter_rows(values_only=True))
                wb2.close()
                if filas2:
                    for fila2 in filas2[1:3]:
                        for celda in fila2:
                            if isinstance(celda, str) and celda.startswith("="):
                                log.info("El Excel contiene fórmulas que podrían no estar cacheadas.")
                                break
            except Exception:
                pass

            log.info("Excel cargado: %s — %d personas, %d columnas", path, len(self.personas), len(self.columnas))
            self._notificar(f"Excel cargado: {len(self.personas)} personas encontradas")

            # Actualizar selector de hoja
            self._actualizar_selector_hoja(sheets_names)

        except Exception as e:
            log.error("Error al leer Excel %s: %s", path, e, exc_info=True)
            messagebox.showerror("Error al leer Excel", str(e))
        finally:
            if ruta_lectura and os.path.exists(ruta_lectura):
                os.unlink(ruta_lectura)

    def _validar_word(self, path: str) -> None:
        try:
            Document(path)
            campos = self._detectar_campos_word()
            if campos:
                log.info("Word validado: %s — %d campos detectados", path, len(campos))
                self._notificar(f"Word cargado. Campos detectados: {', '.join(campos)}")
            else:
                log.info("Word validado: %s — sin campos {{...}} detectados", path)
                self._notificar("Word cargado. No se detectaron campos {{...}}. Asegúrate de usar {{nombre}}, {{rut}}, etc.")
        except Exception as e:
            log.error("Error al leer Word %s: %s", path, e, exc_info=True)
            messagebox.showerror("Error al leer Word", str(e))

    def _generar(self) -> None:
        if not getattr(self, "trabajadores_sel", []):
            messagebox.showwarning("Atencion", "No hay trabajadores seleccionados.")
            return

        carpeta_inicial = self.config_data.get("ultima_carpeta", "")
        carpeta = filedialog.askdirectory(
            title="Selecciona la carpeta donde guardar los documentos",
            initialdir=carpeta_inicial or None,
        )
        if not carpeta:
            return

        self.ultima_carpeta = carpeta
        log.info("Iniciando generación: %d documento(s) en %s", len(self.trabajadores_sel), carpeta)
        self.btn_siguiente.configure(state="disabled", text="Generando...")
        total = len(self.trabajadores_sel)
        self.progress_bar.set(0)
        self.lbl_progreso.configure(text="")
        self.progress_bar.pack(anchor="w", pady=(0, 4))
        self.lbl_progreso.pack(anchor="w")
        threading.Thread(target=self._generar_thread, args=(carpeta,), daemon=True).start()

    def _generar_thread(self, carpeta: str) -> None:
        formato_fecha = getattr(self, "formato_fecha_var", tk.StringVar(value="%d/%m/%Y")).get()
        col_id = getattr(self, "col_identificador_var", tk.StringVar(value=self.columnas[0])).get()
        usar_guion = getattr(self, "espacios_var", tk.StringVar(value="guion_bajo")).get() == "guion_bajo"
        total = len(self.trabajadores_sel)

        def on_progreso(actual: int, total: int) -> None:
            def _update(pct: float, txt: str) -> None:
                self.progress_bar.set(pct)
                self.lbl_progreso.configure(text=txt)
            self.after(0, lambda: _update(actual / total, f"Generando documento {actual} de {total}..."))

        generados, errores = generar_documentos(
            word_path=self.word_path.get(),
            trabajadores=self.trabajadores_sel,
            columnas=self.columnas,
            formato_fecha=formato_fecha,
            col_identificador=col_id,
            campos_extra_vars=self.campos_extra_vars,
            formato_nombre=self.formato_nombre_var.get(),
            usar_guion_bajo=usar_guion,
            carpeta_destino=carpeta,
            generar_pdf=self.pdf_var.get(),
            on_progreso=on_progreso,
        )

        self.after(0, lambda: self._generar_resultado(carpeta, generados, errores))

    def _generar_resultado(self, carpeta: str, generados: int, errores: list[str]) -> None:
        self.btn_siguiente.configure(state="normal", text="Generar", fg_color=VERDE)
        self.progress_bar.pack_forget()
        self.lbl_progreso.pack_forget()
        self.lbl_estado.configure(text=f"{generados} documento(s) generado(s) correctamente.")

        if errores:
            for err in errores:
                log.error("Error en generación: %s", err)
            log_path = os.path.join(carpeta, "_errores.log")
            try:
                with open(log_path, "w", encoding="utf-8") as f:
                    f.write(f"Errores de generación - {datetime.today().strftime('%Y-%m-%d %H:%M:%S')}\n")
                    f.write("=" * 50 + "\n")
                    for err in errores:
                        f.write(err + "\n")
                msg = f"Ocurrieron {len(errores)} error(es).\nSe guardó un registro en:\n{log_path}"
                messagebox.showwarning("Errores durante la generación", msg)
            except Exception as e:
                log.error("No se pudo escribir el log de errores: %s", e)
                messagebox.showwarning("Errores durante la generación", "\n".join(errores[:5]))

        log.info("Generación completada: %d exitosos, %d errores", generados, len(errores))

        if generados > 0:
            respuesta = messagebox.askyesno("Listo", f"Se generaron {generados} documento(s).\n¿Abrir la carpeta?")
            if respuesta:
                if os.name == "nt":
                    os.startfile(carpeta)
                elif sys.platform == "darwin":
                    subprocess.run(["open", carpeta], check=False)
                else:
                    subprocess.run(["xdg-open", carpeta], check=False)

    def _actualizar_preview(self) -> None:
        preview = ""
        if hasattr(self, "trabajadores_sel") and self.trabajadores_sel:
            persona = self.trabajadores_sel[0]
            formato = self.formato_nombre_var.get()
            try:
                for k, v in persona.items():
                    formato = formato.replace(f"{{{k}}}", str(v) if v is not None else "")
                preview = formato.replace("{fecha}", datetime.today().strftime("%Y%m%d"))
                usar_guion = getattr(self, "espacios_var", tk.StringVar(value="guion_bajo")).get() == "guion_bajo"
                preview = limpiar_nombre(preview, usar_guion)
                preview = preview + ".docx" if preview else ""
            except Exception:
                preview = ""
        self.lbl_preview.configure(text=f"Ejemplo: {preview}" if preview else "")

    def _guardar_config(self) -> None:
        self.config_data["ancho"] = self.winfo_width()
        self.config_data["alto"] = self.winfo_height()
        if hasattr(self, "formato_nombre_var"):
            self.config_data["formato_nombre"] = self.formato_nombre_var.get()
        if hasattr(self, "col_identificador_var"):
            self.config_data["col_identificador"] = self.col_identificador_var.get()
        if hasattr(self, "espacios_var"):
            self.config_data["usar_guion_bajo"] = self.espacios_var.get() == "guion_bajo"
        if hasattr(self, "formato_fecha_var"):
            self.config_data["formato_fecha"] = self.formato_fecha_var.get()
        if hasattr(self, "ultima_carpeta"):
            self.config_data["ultima_carpeta"] = self.ultima_carpeta
        guardar(self.config_data)

    def _actualizar_selector_hoja(self, sheets: list[str]) -> None:
        if len(sheets) > 1:
            actual = self.sheet_var.get()
            self.hoja_combo.configure(values=sheets)
            if actual in sheets:
                self.sheet_var.set(actual)
            else:
                self.sheet_var.set(sheets[0])
            self.frame_hoja.pack(fill="x", pady=(0, 10), before=self.frame_hoja.master.winfo_children()[-1])
        else:
            self.frame_hoja.pack_forget()

    def _on_hoja_cambiada(self, _: str) -> None:
        self._cargar_excel(self.excel_path.get())

    def _notificar(self, msg: str) -> None:
        original = self.title()
        self.title(msg)
        self.after(3000, lambda: self.title(original))


if __name__ == "__main__":
    app = AppContratos()
    app.mainloop()
